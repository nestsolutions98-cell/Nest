from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from sqlalchemy.exc import IntegrityError
import os
import logging
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import requests

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Use an instance folder for writable files (SQLite DB, secrets, uploads)
app = Flask(__name__, instance_relative_config=True)

# Make sure the instance/ directory exists
os.makedirs(app.instance_path, exist_ok=True)

# Database configuration - keep local data separate from server data
# Treat local run (dev server) as SQLite even if DATABASE_URL exists.
IS_LOCAL_RUN = (
    os.environ.get('FLASK_ENV') == 'development' or
    not os.getenv('GUNICORN_CMD_ARGS') or                 # usually absent when running `python app.py`
    os.getenv('RUN_LOCAL') == '1'                         # explicit override
)

if not IS_LOCAL_RUN and os.getenv('DATABASE_URL'):
    # Remote/production: use Postgres (e.g., Railway) via DATABASE_URL
    app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
else:
    # Local development: use an isolated SQLite file
    db_path = os.path.join(app.instance_path, 'sport_courses_local.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{db_path}"

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', os.urandom(24).hex())
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
CORS(app)

# ---- Simple session-based auth configuration ----
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME', 'admin')
# Prefer a password hash via ADMIN_PASSWORD_HASH, else hash ADMIN_PASSWORD at startup
from werkzeug.security import check_password_hash, generate_password_hash

_pwd_hash_env = os.getenv('ADMIN_PASSWORD_HASH')
_pwd_plain = os.getenv('ADMIN_PASSWORD', 'admin123')
ADMIN_PASSWORD_HASH = _pwd_hash_env or generate_password_hash(_pwd_plain)


def login_required(view_func):
    from functools import wraps
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get('user'):
            next_url = request.path
            return redirect(url_for('login', next=next_url))
        return view_func(*args, **kwargs)

    return wrapped

# ---------------- Green Invoice config ----------------
GI_API_BASE = os.getenv('GI_API_BASE', 'https://api.greeninvoice.co.il/api/v1')
GI_CLIENT_ID = os.getenv('GI_CLIENT_ID')
GI_CLIENT_SECRET = os.getenv('GI_CLIENT_SECRET')
try:
    GI_DOC_TYPE = int(os.getenv('GI_DOC_TYPE', '400'))  # 400 invoice/receipt; 320 tax invoice
except Exception:
    GI_DOC_TYPE = 400
GI_SEND_EMAIL = (os.getenv('GI_SEND_EMAIL', 'false').lower() == 'true')

def _gi_token():
    if not GI_CLIENT_ID or not GI_CLIENT_SECRET:
        raise RuntimeError('Green Invoice credentials are not configured')
    r = requests.post(f'{GI_API_BASE}/account/token', json={'id': GI_CLIENT_ID, 'secret': GI_CLIENT_SECRET}, timeout=20)
    r.raise_for_status()
    data = r.json()
    token = data.get('token')
    if not token:
        raise RuntimeError('Failed to fetch Green Invoice token')
    return token

def _gi_headers(token: str):
    return {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}


# Database Models
class Course(db.Model):
    __tablename__ = 'courses'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    teacher = db.Column(db.String(100), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    time = db.Column(db.String(10), nullable=False)  # Format: "HH:MM"
    duration = db.Column(db.Integer, nullable=False, default=60)  # Duration in minutes for each meeting
    sessions_count = db.Column(db.Integer, nullable=False)
    sessions_per_week = db.Column(db.Integer, nullable=False)
    weekdays = db.Column(db.String(50), nullable=False)  # Format: "0,2,5" (Sunday=0, Saturday=6)
    end_date = db.Column(db.Date, nullable=False)
    color = db.Column(db.String(7), nullable=False, default='#3B82F6')  # Default blue color

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'teacher': self.teacher,
            'start_date': self.start_date.strftime('%Y-%m-%d'),
            'time': self.time,
            'duration': self.duration,
            'sessions_count': self.sessions_count,
            'sessions_per_week': self.sessions_per_week,
            'weekdays': self.weekdays,
            'end_date': self.end_date.strftime('%Y-%m-%d'),
            'color': self.color
        }


class Coach(db.Model):
    __tablename__ = 'coaches'
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    phone = db.Column(db.String(20), unique=True, nullable=False)

    def to_dict(self):
        return {
            'id': self.id,
            'first_name': self.first_name,
            'last_name': self.last_name,
            'phone': self.phone,
            'full_name': f"{self.first_name} {self.last_name}".strip()
        }

class Student(db.Model):
    __tablename__ = 'students'
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(50), nullable=False)
    fathers_name = db.Column(db.String(50), nullable=False)
    phone = db.Column(db.String(20), unique=True, nullable=False)
    date_of_birth = db.Column(db.Date, nullable=False)
    national_id = db.Column(db.String(20), unique=True, nullable=True)
    enrollments = db.relationship('Enrollment', backref='student', cascade='all, delete-orphan')

    def to_dict(self):
        return {
            'id': self.id,
            'first_name': self.first_name,
            'fathers_name': self.fathers_name,
            'phone': self.phone,
            'date_of_birth': self.date_of_birth.strftime('%Y-%m-%d'),
            'national_id': self.national_id,
            'age': self.calculate_age()
        }

    def calculate_age(self):
        today = datetime.now().date()
        return today.year - self.date_of_birth.year - (
                (today.month, today.day) < (self.date_of_birth.month, self.date_of_birth.day))


class Enrollment(db.Model):
    __tablename__ = 'course_enrollments'
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('courses.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    enrollment_date = db.Column(db.Date, default=datetime.now().date)
    course = db.relationship('Course', backref='enrollments')


class Payment(db.Model):
    __tablename__ = 'payments'
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    course_id = db.Column(db.Integer, db.ForeignKey('courses.id'), nullable=False)
    month = db.Column(db.String(7), nullable=False)  # Format: "YYYY-MM"
    amount = db.Column(db.Float, nullable=False, default=0.0)
    payment_date = db.Column(db.Date, default=datetime.now().date)
    payment_method = db.Column(db.String(20), nullable=True)  # cash, check, transfer, etc.

    student = db.relationship('Student', backref='payments')
    course = db.relationship('Course', backref='payments')

    def to_dict(self):
        return {
            'id': self.id,
            'student_id': self.student_id,
            'course_id': self.course_id,
            'student_name': f"{self.student.first_name} {self.student.fathers_name}",
            'course_name': self.course.name,
            'teacher_name': self.course.teacher,
            'month': self.month,
            'amount': self.amount,
            'payment_date': self.payment_date.strftime('%Y-%m-%d'),
            'payment_method': self.payment_method
        }


class CourseMeeting(db.Model):
    __tablename__ = 'course_meetings'
    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey('courses.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    notes = db.Column(db.String(255))
    course = db.relationship('Course', backref=db.backref('meetings', lazy=True))
    attendances = db.relationship('Attendance', backref='meeting', lazy=True)


class Attendance(db.Model):
    __tablename__ = 'attendances'
    id = db.Column(db.Integer, primary_key=True)
    meeting_id = db.Column(db.Integer, db.ForeignKey('course_meetings.id'), nullable=False)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    present = db.Column(db.Boolean, default=False)
    student = db.relationship('Student', backref=db.backref('attendances', lazy=True))


def create_database():
    with app.app_context():
        db.create_all()

        # Add color column to existing courses if it doesn't exist
        try:
            if 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']:
                db.session.execute(text('ALTER TABLE courses ADD COLUMN color VARCHAR(7) DEFAULT "#3B82F6"'))
            else:
                db.session.execute(text('ALTER TABLE courses ADD COLUMN color VARCHAR(7) DEFAULT \'#3B82F6\''))
            db.session.commit()
            logger.info("Added color column to existing courses")
        except Exception as e:
            # Column might already exist
            logger.info(f"Color column might already exist: {e}")

        logger.info("Database created successfully with correct schema!")


def reset_database():
    """Reset the database completely - useful for Railway deployment"""
    with app.app_context():
        try:
            # Drop all tables
            db.drop_all()
            logger.info("Dropped all existing tables")
            
            # Create all tables with current schema
            db.create_all()
            logger.info("Created all tables with current schema")
            
            return True
        except Exception as e:
            logger.error(f"Error resetting database: {e}")
            return False


def migrate_existing_data():
    with app.app_context():
        # Check if we need to migrate existing tables
        inspector = db.inspect(db.engine)
        table_names = inspector.get_table_names()
        
        # If no tables exist, create all tables
        if not table_names:
            db.create_all()
            logger.info("Database created successfully with correct schema!")
            return
            
        # Check if courses table exists and has all required columns
        if 'courses' in table_names:
            course_columns = [col['name'] for col in inspector.get_columns('courses')]
            missing_columns = []
            
            if 'duration' not in course_columns:
                missing_columns.append('duration')
            if 'color' not in course_columns:
                missing_columns.append('color')
                
            # Add missing columns based on database type
            if missing_columns:
                try:
                    if 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']:
                        # SQLite syntax
                        for col in missing_columns:
                            if col == 'duration':
                                db.session.execute(text('ALTER TABLE courses ADD COLUMN duration INTEGER DEFAULT 60'))
                            elif col == 'color':
                                db.session.execute(text('ALTER TABLE courses ADD COLUMN color VARCHAR(7) DEFAULT "#3B82F6"'))
                    else:
                        # PostgreSQL syntax
                        for col in missing_columns:
                            if col == 'duration':
                                db.session.execute(text('ALTER TABLE courses ADD COLUMN duration INTEGER DEFAULT 60'))
                            elif col == 'color':
                                db.session.execute(text('ALTER TABLE courses ADD COLUMN color VARCHAR(7) DEFAULT \'#3B82F6\''))
                    
                    db.session.commit()
                    logger.info(f"Added missing columns to courses table: {missing_columns}")
                except Exception as e:
                    logger.error(f"Error adding columns to courses table: {e}")
                    # If ALTER TABLE fails, recreate the table
                    try:
                        db.session.execute(text('DROP TABLE IF EXISTS courses CASCADE'))
                        db.create_all()
                        logger.info("Recreated courses table with correct schema")
                    except Exception as e2:
                        logger.error(f"Error recreating courses table: {e2}")
        
        # Check payments table
        if 'payments' in table_names:
            payment_columns = [col['name'] for col in inspector.get_columns('payments')]
            missing_payment_columns = []
            
            if 'course_id' not in payment_columns or 'amount' not in payment_columns:
                # Drop and recreate the payments table
                try:
                    db.session.execute(text('DROP TABLE IF EXISTS payments CASCADE'))
                    db.create_all()
                    logger.info("Payment table recreated successfully!")
                except Exception as e:
                    logger.error(f"Error recreating payments table: {e}")
            elif 'payment_method' not in payment_columns:
                # Add payment_method column
                try:
                    if 'sqlite' in app.config['SQLALCHEMY_DATABASE_URI']:
                        db.session.execute(text('ALTER TABLE payments ADD COLUMN payment_method VARCHAR(20)'))
                    else:
                        db.session.execute(text('ALTER TABLE payments ADD COLUMN payment_method VARCHAR(20)'))
                    db.session.commit()
                    logger.info("Added payment_method column to existing payments table")
                except Exception as e:
                    logger.error(f"Error adding payment_method column: {e}")
        
        # Ensure all tables exist
        db.create_all()

        # Ensure coaches table exists for existing DBs without running full migrations in prod
        try:
            inspector = db.inspect(db.engine)
            if 'coaches' not in inspector.get_table_names():
                Coach.__table__.create(db.engine)
                logger.info("Created missing 'coaches' table")
        except Exception as e:
            logger.error(f"Error ensuring coaches table exists: {e}")
        logger.info("Database migration completed successfully!")


def compute_schedule_metrics(weekdays_str: str, sessions_count: int):
    """
    Return (sessions_per_week, total_weeks) derived from weekdays_str and sessions_count.
    weekdays_str is a comma-separated list like "0,2,5".
    Raises ValueError if no weekdays are selected.
    """
    weekdays_list = [d for d in (weekdays_str or "").split(",") if d.strip() != ""]
    sessions_per_week = len(set(weekdays_list))
    if sessions_per_week == 0:
        raise ValueError("At least one weekday must be selected")

    # ceil division
    total_weeks = (sessions_count + sessions_per_week - 1) // sessions_per_week
    return sessions_per_week, total_weeks


# Routes
@app.route('/')
@login_required
def home():
    return render_template('index.html')


# --------------- Auth routes ---------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        next_url = request.args.get('next') or request.form.get('next') or url_for('home')

        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session['user'] = username
            return redirect(next_url)
        else:
            # Render with error
            return render_template('login.html', error='Invalid credentials', next_url=next_url), 401

    # GET
    next_url = request.args.get('next', url_for('home'))
    return render_template('login.html', next_url=next_url)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/admin/reset-database', methods=['POST'])
def admin_reset_database():
    """Reset database - only accessible in development or with proper auth"""
    # Only allow in development or if user is admin
    if os.getenv('FLASK_ENV') != 'development' and not session.get('user'):
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        if reset_database():
            return jsonify({'message': 'Database reset successfully'}), 200
        else:
            return jsonify({'error': 'Failed to reset database'}), 500
    except Exception as e:
        logger.error(f"Error in admin reset database: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/course/<int:course_id>')
def course_profile(course_id):
    # Protect with login
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    course = Course.query.get_or_404(course_id)
    # Get enrolled students
    enrollments = Enrollment.query.filter_by(course_id=course_id).all()
    students = [enrollment.student.to_dict() for enrollment in enrollments]
    # Get meetings (with attendance)
    meetings = CourseMeeting.query.filter_by(course_id=course_id).order_by(CourseMeeting.date.desc()).all()
    meetings_data = []
    for meeting in meetings:
        attendance = Attendance.query.filter_by(meeting_id=meeting.id).all()
        attendance_list = [
            {
                'student_id': a.student_id,
                'student_name': a.student.first_name + ' ' + a.student.fathers_name,
                'present': a.present
            } for a in attendance
        ]
        meetings_data.append({
            'id': meeting.id,
            'date': meeting.date.strftime('%Y-%m-%d'),
            'notes': meeting.notes,
            'attendance': attendance_list
        })
    # Calculate meetings left
    total_meetings = course.sessions_count
    meetings_left = max(0, total_meetings - len(meetings))
    return render_template(
        'course_profile.html',
        course_id=course.id,
        course_name=course.name,
        course=course,
        students=students,
        meetings=meetings_data,
        meetings_left=meetings_left
    )


@app.route('/api/courses')
def get_courses():
    courses = Course.query.all()
    return jsonify([course.to_dict() for course in courses])


@app.route('/api/courses', methods=['POST'])
def create_course():
    data = request.json or {}

    # Defaults + required fields
    color = data.get('color', '#3B82F6')
    start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    sessions_count = int(data['sessions_count'])
    duration = int(data.get('duration', 60))
    weekdays_str = data.get('weekdays', '')

    # Derive sessions_per_week and number of weeks
    try:
        sessions_per_week, total_weeks = compute_schedule_metrics(weekdays_str, sessions_count)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    end_date = start_date + timedelta(weeks=total_weeks)

    course = Course(
        name=data['name'],
        teacher=data['teacher'],
        start_date=start_date,
        time=data['time'],
        duration=duration,
        sessions_count=sessions_count,
        sessions_per_week=sessions_per_week,  # stored but auto-computed
        weekdays=weekdays_str,
        end_date=end_date,
        color=color
    )
    try:
        db.session.add(course)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': 'Course name must be unique'}), 409
    return jsonify(course.to_dict()), 201


@app.route('/api/courses/<int:course_id>', methods=['DELETE'])
def delete_course(course_id):
    try:
        course = Course.query.get_or_404(course_id)

        # Check if course can be safely deleted
        enrollment_count = Enrollment.query.filter_by(course_id=course_id).count()
        payment_count = Payment.query.filter_by(course_id=course_id).count()
        meeting_count = CourseMeeting.query.filter_by(course_id=course_id).count()

        if enrollment_count > 0 or payment_count > 0 or meeting_count > 0:
            logger.info(
                f"Course {course_id} has related data: {enrollment_count} enrollments, {payment_count} payments, {meeting_count} meetings")

        # Delete all related data first
        # 1. Delete all enrollments for this course
        Enrollment.query.filter_by(course_id=course_id).delete()

        # 2. Delete all payments for this course
        Payment.query.filter_by(course_id=course_id).delete()

        # 3. Delete all course meetings and attendance
        meetings = CourseMeeting.query.filter_by(course_id=course_id).all()
        for meeting in meetings:
            # Delete attendance records for this meeting
            Attendance.query.filter_by(meeting_id=meeting.id).delete()
            # Delete the meeting
            db.session.delete(meeting)

        # 4. Now delete the course
        db.session.delete(course)
        db.session.commit()

        logger.info(f"Course {course_id} deleted successfully")
        return '', 204

    except Exception as e:
        logger.error(f"Error deleting course {course_id}: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Error deleting course: {str(e)}'}), 500


@app.route('/api/courses/<int:course_id>', methods=['PUT'])
def update_course(course_id):
    data = request.json or {}
    course = Course.query.get_or_404(course_id)

    # Basic fields
    course.name = data.get('name', course.name)
    course.teacher = data.get('teacher', course.teacher)
    course.time = data.get('time', course.time)
    course.color = data.get('color', course.color)

    # Recompute schedule info from weekdays + sessions_count
    sessions_count = int(data.get('sessions_count', course.sessions_count))
    weekdays_str = data.get('weekdays', course.weekdays)

    try:
        sessions_per_week, total_weeks = compute_schedule_metrics(weekdays_str, sessions_count)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    course.sessions_count = sessions_count
    course.weekdays = weekdays_str
    course.sessions_per_week = sessions_per_week
    course.end_date = course.start_date + timedelta(weeks=total_weeks)

    db.session.commit()
    return jsonify(course.to_dict())


# ---------------- Coaches API ----------------
@app.route('/api/coaches')
def get_coaches():
    coaches = Coach.query.all()
    return jsonify([c.to_dict() for c in coaches])


@app.route('/api/coaches/<int:coach_id>', methods=['GET'])
def get_coach(coach_id):
    coach = Coach.query.get_or_404(coach_id)
    return jsonify(coach.to_dict())


@app.route('/api/coaches', methods=['POST'])
def create_coach():
    data = request.json or {}
    try:
        coach = Coach(
            first_name=data.get('first_name', '').strip(),
            last_name=data.get('last_name', '').strip(),
            phone=data.get('phone', '').strip(),
        )
        if not coach.first_name or not coach.last_name or not coach.phone:
            return jsonify({'error': 'first_name, last_name and phone are required'}), 400
        db.session.add(coach)
        db.session.commit()
        return jsonify(coach.to_dict()), 201
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': 'Coach phone must be unique'}), 409


@app.route('/api/coaches/<int:coach_id>', methods=['PUT'])
def update_coach(coach_id):
    data = request.json or {}
    coach = Coach.query.get_or_404(coach_id)
    coach.first_name = data.get('first_name', coach.first_name)
    coach.last_name = data.get('last_name', coach.last_name)
    coach.phone = data.get('phone', coach.phone)
    try:
        db.session.commit()
        return jsonify(coach.to_dict())
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': 'Coach phone must be unique'}), 409


@app.route('/api/coaches/<int:coach_id>', methods=['DELETE'])
def delete_coach(coach_id):
    coach = Coach.query.get_or_404(coach_id)
    db.session.delete(coach)
    db.session.commit()
    return '', 204


@app.route('/coach/<int:coach_id>')
def coach_profile(coach_id):
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))

    coach = Coach.query.get_or_404(coach_id)
    full_name = f"{coach.first_name} {coach.last_name}".strip()

    # Find courses taught by this coach by matching the teacher field to full name
    courses = Course.query.filter(Course.teacher == full_name).all()

    # Prepare course dicts for template
    courses_data = [c.to_dict() for c in courses]

    return render_template('coach.html', coach=coach.to_dict(), courses=courses_data)


@app.route('/api/courses/<int:course_id>')
def get_course(course_id):
    course = Course.query.get_or_404(course_id)
    return jsonify(course.to_dict())


@app.route('/api/courses/<int:course_id>/students')
def get_course_students(course_id):
    enrollments = Enrollment.query.filter_by(course_id=course_id).all()
    students = []
    for enrollment in enrollments:
        student_data = enrollment.student.to_dict()
        student_data['enrollment_id'] = enrollment.id
        students.append(student_data)
    return jsonify(students)


@app.route('/api/students')
def get_students():
    students = Student.query.all()
    return jsonify([student.to_dict() for student in students])


@app.route('/api/students', methods=['POST'])
def create_student():
    data = request.json
    date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()

    student = Student(
        first_name=data['first_name'],
        fathers_name=data['fathers_name'],
        phone=data['phone'],
        date_of_birth=datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date(),
        national_id=(data.get('national_id') or None)
    )
    try:
        db.session.add(student)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': 'Student phone or national ID must be unique'}), 409
    return jsonify(student.to_dict()), 201


@app.route('/api/students/<int:student_id>', methods=['DELETE'])
def delete_student(student_id):
    try:
        student = Student.query.get_or_404(student_id)

        # Check if student can be safely deleted
        enrollment_count = Enrollment.query.filter_by(student_id=student_id).count()
        payment_count = Payment.query.filter_by(student_id=student_id).count()
        attendance_count = Attendance.query.filter_by(student_id=student_id).count()

        if enrollment_count > 0 or payment_count > 0 or attendance_count > 0:
            logger.info(
                f"Student {student_id} has related data: {enrollment_count} enrollments, {payment_count} payments, {attendance_count} attendance records")

        # Delete all related data first
        # 1. Delete all enrollments for this student
        Enrollment.query.filter_by(student_id=student_id).delete()

        # 2. Delete all payments for this student
        Payment.query.filter_by(student_id=student_id).delete()

        # 3. Delete all attendance records for this student
        Attendance.query.filter_by(student_id=student_id).delete()

        # 4. Now delete the student
        db.session.delete(student)
        db.session.commit()

        logger.info(f"Student {student_id} deleted successfully")
        return '', 204

    except Exception as e:
        logger.error(f"Error deleting student {student_id}: {str(e)}")
        db.session.rollback()
        return jsonify({'error': f'Error deleting student: {str(e)}'}), 500


@app.route('/api/students/<int:student_id>', methods=['PUT'])
def update_student(student_id):
    data = request.json
    student = Student.query.get_or_404(student_id)
    student.first_name = data.get('first_name', student.first_name)
    student.fathers_name = data.get('fathers_name', student.fathers_name)
    student.phone = data.get('phone', student.phone)
    if 'date_of_birth' in data:
        student.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
    student.national_id = (data.get('national_id') if data.get('national_id') not in [None, '', ' '] else None)
    db.session.commit()
    return jsonify(student.to_dict())


@app.route('/api/students/<int:student_id>/payments')
def get_student_payments(student_id):
    payments = Payment.query.filter_by(student_id=student_id).all()
    return jsonify(
        [{'month': payment.month, 'payment_date': payment.payment_date.strftime('%Y-%m-%d')} for payment in payments])


@app.route('/api/students/<int:student_id>/payments', methods=['POST'])
def add_payment(student_id):
    data = request.json
    # Normalize payment method; default to 'cash' if missing/invalid
    method = (data.get('payment_method') or '').strip().lower()
    if method not in ('cash', 'check', 'transfer'):
        method = 'cash'

    payment = Payment(
        student_id=student_id,
        month=data['month'],
        course_id=data.get('course_id'),  # Make course_id optional for backward compatibility
        amount=data.get('amount', 0.0),  # Make amount optional for backward compatibility
        payment_method=method  # normalized
    )
    db.session.add(payment)
    db.session.commit()
    return jsonify({'month': payment.month}), 201


@app.route('/api/students/<int:student_id>/courses')
def get_student_courses(student_id):
    enrollments = Enrollment.query.filter_by(student_id=student_id).all()
    courses = [enrollment.course.to_dict() for enrollment in enrollments]
    return jsonify(courses)


@app.route('/api/enrollments', methods=['POST'])
def enroll_student():
    data = request.json
    enrollment = Enrollment(
        course_id=data['course_id'],
        student_id=data['student_id']
    )
    db.session.add(enrollment)
    db.session.commit()
    return jsonify({'id': enrollment.id}), 201


@app.route('/api/enrollments/<int:enrollment_id>', methods=['DELETE'])
def remove_enrollment(enrollment_id):
    enrollment = Enrollment.query.get_or_404(enrollment_id)
    db.session.delete(enrollment)
    db.session.commit()
    return '', 204


@app.route('/api/enrollments')
def get_enrollments():
    enrollments = Enrollment.query.all()
    return jsonify([{'course_id': e.course_id, 'student_id': e.student_id} for e in enrollments])


@app.route('/api/calendar/weekly')
def get_weekly_calendar():
    start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-%d'))
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = start_date + timedelta(days=6)

    courses = Course.query.filter(
        Course.start_date <= end_date,
        Course.end_date >= start_date
    ).all()

    calendar_data = []
    for course in courses:
        # Parse weekdays (frontend sends Sunday-based: 0=Sunday, 1=Monday, ..., 6=Saturday)
        weekdays = [int(day) for day in course.weekdays.split(',')]

        # Generate course sessions for the week based on weekdays
        current_date = max(course.start_date, start_date)
        while current_date <= min(course.end_date, end_date):
            # Get the weekday in Sunday-based format (0=Sunday, 6=Saturday)
            # Python weekday(): Monday=0, Sunday=6
            # We need: Sunday=0, Saturday=6
            weekday = current_date.weekday()
            sunday_based_weekday = (weekday + 1) % 7

            if sunday_based_weekday in weekdays:
                # Calculate classes remaining correctly
                # Parse weekdays to get the specific days this course runs on
                course_weekdays = [int(day) for day in course.weekdays.split(',')]

                # Count how many classes have already occurred
                classes_completed = 0
                temp_date = course.start_date

                # Count classes from start date up to (but not including) current date
                while temp_date < current_date:
                    # Convert to Sunday-based weekday (0=Sunday, 1=Monday, etc.)
                    temp_weekday = (temp_date.weekday() + 1) % 7  # Convert Monday=0 to Sunday=0
                    if temp_weekday in course_weekdays:
                        classes_completed += 1
                    temp_date += timedelta(days=1)

                classes_remaining = max(0, course.sessions_count - classes_completed)

                # Get number of enrolled students
                enrolled_count = db.session.query(Enrollment).filter_by(course_id=course.id).count()

                calendar_data.append({
                    'id': course.id,
                    'title': course.name,
                    'date': current_date.strftime('%Y-%m-%d'),
                    'time': course.time,
                    'duration': course.duration,
                    'teacher': course.teacher,
                    'color': course.color,
                    'enrolled_count': enrolled_count,
                    'classes_remaining': classes_remaining
                })
            current_date += timedelta(days=1)

    # Sort calendar data by date and time
    calendar_data.sort(key=lambda x: (x['date'], x['time']))

    return jsonify(calendar_data)


@app.route('/api/calendar/daily')
def get_daily_calendar():
    """Return calendar events for a single day."""
    date_str = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    target_date = datetime.strptime(date_str, '%Y-%m-%d').date()

    courses = Course.query.filter(
        Course.start_date <= target_date,
        Course.end_date >= target_date
    ).all()

    calendar_data = []
    for course in courses:
        weekdays = [int(day) for day in course.weekdays.split(',') if day != '']
        weekday = (target_date.weekday() + 1) % 7  # Convert Monday=0 to Sunday=0

        if weekday in weekdays:
            # Calculate remaining classes up to the target date
            course_weekdays = weekdays
            classes_completed = 0
            temp_date = course.start_date
            while temp_date < target_date:
                temp_weekday = (temp_date.weekday() + 1) % 7
                if temp_weekday in course_weekdays:
                    classes_completed += 1
                temp_date += timedelta(days=1)

            classes_remaining = max(0, course.sessions_count - classes_completed)
            enrolled_count = db.session.query(Enrollment).filter_by(course_id=course.id).count()

            calendar_data.append({
                'id': course.id,
                'title': course.name,
                'date': target_date.strftime('%Y-%m-%d'),
                'time': course.time,
                'duration': course.duration,
                'teacher': course.teacher,
                'color': course.color,
                'enrolled_count': enrolled_count,
                'classes_remaining': classes_remaining
            })

    calendar_data.sort(key=lambda x: x['time'])
    return jsonify(calendar_data)


@app.route('/api/whatsapp/send')
def send_whatsapp():
    phone = request.args.get('phone')
    message = request.args.get('message', '')
    whatsapp_url = f"https://wa.me/{phone}?text={message}"
    return jsonify({'url': whatsapp_url})


@app.route('/api/payments')
def get_payments():
    payments = Payment.query.all()
    return jsonify([payment.to_dict() for payment in payments])


@app.route('/api/payments', methods=['POST'])
def create_payment():
    data = request.json
    # Normalize payment method; default to 'cash' if missing/invalid
    method = (data.get('payment_method') or '').strip().lower()
    if method not in ('cash', 'check', 'transfer'):
        method = 'cash'

    payment = Payment(
        student_id=data['student_id'],
        course_id=data['course_id'],
        month=data['month'],
        amount=data['amount'],
        payment_method=method  # normalized
    )
    db.session.add(payment)
    db.session.commit()
    return jsonify(payment.to_dict()), 201


@app.route('/api/payments/<int:payment_id>', methods=['DELETE'])
def delete_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    db.session.delete(payment)
    db.session.commit()
    return '', 204


@app.route('/api/analysis/coach-income')
def get_coach_income():
    period = request.args.get('period', 'all')

    # Build date filter based on period
    if period == 'month':
        start_date = datetime.now().replace(day=1).date()
    elif period == 'quarter':
        current_month = datetime.now().month
        quarter_start_month = ((current_month - 1) // 3) * 3 + 1
        start_date = datetime.now().replace(month=quarter_start_month, day=1).date()
    elif period == 'year':
        start_date = datetime.now().replace(month=1, day=1).date()
    else:  # all
        start_date = None

    # Query payments with course and teacher info
    query = db.session.query(
        Course.teacher,
        db.func.sum(Payment.amount).label('total_income'),
        db.func.count(Payment.id).label('payment_count')
    ).join(Payment.course).group_by(Course.teacher)

    if start_date:
        query = query.filter(Payment.payment_date >= start_date)

    results = query.all()

    return jsonify([{
        'teacher': result.teacher,
        'total_income': float(result.total_income or 0),
        'payment_count': result.payment_count
    } for result in results])


@app.route('/api/analysis/course-income')
def get_course_income():
    period = request.args.get('period', 'all')

    # Build date filter based on period
    if period == 'month':
        start_date = datetime.now().replace(day=1).date()
    elif period == 'quarter':
        current_month = datetime.now().month
        quarter_start_month = ((current_month - 1) // 3) * 3 + 1
        start_date = datetime.now().replace(month=quarter_start_month, day=1).date()
    elif period == 'year':
        start_date = datetime.now().replace(month=1, day=1).date()
    else:  # all
        start_date = None

    # Query payments with course info
    query = db.session.query(
        Course.name,
        db.func.sum(Payment.amount).label('total_income'),
        db.func.count(Payment.id).label('payment_count')
    ).join(Payment.course).group_by(Course.name)

    if start_date:
        query = query.filter(Payment.payment_date >= start_date)

    results = query.all()

    return jsonify([{
        'course_name': result.name,
        'total_income': float(result.total_income or 0),
        'payment_count': result.payment_count
    } for result in results])


@app.route('/api/analysis/summary')
def get_analysis_summary():
    period = request.args.get('period', 'all')

    # Build date filter based on period
    if period == 'month':
        start_date = datetime.now().replace(day=1).date()
    elif period == 'quarter':
        current_month = datetime.now().month
        quarter_start_month = ((current_month - 1) // 3) * 3 + 1
        start_date = datetime.now().replace(month=quarter_start_month, day=1).date()
    elif period == 'year':
        start_date = datetime.now().replace(month=1, day=1).date()
    else:  # all
        start_date = None

    # Query summary statistics
    query = db.session.query(
        db.func.sum(Payment.amount).label('total_revenue'),
        db.func.count(Payment.id).label('total_payments'),
        db.func.avg(Payment.amount).label('average_payment')
    )

    if start_date:
        query = query.filter(Payment.payment_date >= start_date)

    result = query.first()

    return jsonify({
        'total_revenue': float(result.total_revenue or 0),
        'total_payments': result.total_payments or 0,
        'average_payment': float(result.average_payment or 0)
    })


@app.route('/api/invoice/<int:payment_id>')
def generate_invoice(payment_id):
    try:
        payment = Payment.query.get_or_404(payment_id)

        # Check if relationships exist
        if not payment.student:
            return jsonify({'error': 'Student not found for this payment'}), 404

        if not payment.course:
            return jsonify({'error': 'Course not found for this payment'}), 404

        # Generate a simple invoice (in a real app, you'd use a proper PDF library)
        invoice_data = {
            'invoice_number': f"INV-{payment.id:06d}",
            'payment_date': payment.payment_date.strftime('%Y-%m-%d'),
            'student_name': f"{payment.student.first_name} {payment.student.fathers_name}",
            'course_name': payment.course.name,
            'teacher_name': payment.course.teacher,
            'month': payment.month,
            'amount': payment.amount,
            'amount_text': f"₪{payment.amount:.2f}",
            'payment_method': payment.payment_method
        }

        return jsonify(invoice_data)
    except Exception as e:
        logger.error(f"Error generating invoice for payment {payment_id}: {str(e)}")
        return jsonify({'error': f'Error generating invoice: {str(e)}'}), 500


@app.route('/api/green-invoice/<int:payment_id>', methods=['POST'])
def create_green_invoice(payment_id: int):
    """Create a real invoice in Green Invoice for the given payment.
    Requires GI_CLIENT_ID and GI_CLIENT_SECRET to be configured.
    """
    try:
        payment = Payment.query.get_or_404(payment_id)

        if not payment.student:
            return jsonify({'error': 'Student not found for this payment'}), 404
        if not payment.course:
            return jsonify({'error': 'Course not found for this payment'}), 404

        token = _gi_token()

        student_name = f"{payment.student.first_name} {payment.student.fathers_name}".strip()
        course_desc = f"{payment.course.name} - {payment.month}"

        payload = {
            'type': GI_DOC_TYPE,
            'lang': 'he',
            'currency': 'ILS',
            'income': True,
            'rounding': 0,
            'client': {
                'name': student_name,
                'phones': [payment.student.phone] if payment.student.phone else [],
                'country': 'IL'
            },
            'items': [{
                'description': course_desc,
                'quantity': 1,
                'price': float(payment.amount)
                # 'vatType': 0  # If VAT included; align with your GI settings
            }],
            'sendEmail': GI_SEND_EMAIL
        }

        r = requests.post(f'{GI_API_BASE}/documents', headers=_gi_headers(token), json=payload, timeout=30)
        if not r.ok:
            return jsonify({'error': 'green_invoice_create_failed', 'status': r.status_code, 'details': r.text}), r.status_code
        doc = r.json() or {}
        doc_id = doc.get('id') or doc.get('_id')

        # Try issuing (some accounts require it)
        try:
            requests.post(f'{GI_API_BASE}/documents/issue', headers=_gi_headers(token), json={'ids': [doc_id]}, timeout=20)
        except Exception:
            pass

        # Direct URL (if present)
        url = doc.get('url')
        if url:
            return jsonify({'ok': True, 'docId': doc_id, 'url': url, 'doc': doc})

        # Try fetching PDF (base64) as a fallback
        pdf_data = None
        try:
            pr = requests.get(f'{GI_API_BASE}/documents/{doc_id}/pdf', headers=_gi_headers(token), timeout=30)
            if pr.ok and pr.headers.get('Content-Type', '').startswith('application/json'):
                pdf_data = pr.json().get('data')
        except Exception:
            pass

        return jsonify({'ok': True, 'docId': doc_id, 'pdf': pdf_data, 'doc': doc})
    except Exception as e:
        logger.error(f"Error creating Green Invoice for payment {payment_id}: {e}")
        return jsonify({'error': 'green_invoice_error', 'message': str(e)}), 500


@app.route('/api/courses/<int:course_id>/meetings', methods=['GET'])
def get_course_meetings(course_id):
    meetings = CourseMeeting.query.filter_by(course_id=course_id).order_by(CourseMeeting.date.desc()).all()
    result = []
    for meeting in meetings:
        attendance = Attendance.query.filter_by(meeting_id=meeting.id).all()
        attendance_list = [
            {
                'student_id': a.student_id,
                'student_name': a.student.first_name + ' ' + a.student.fathers_name,
                'present': a.present
            } for a in attendance
        ]
        result.append({
            'id': meeting.id,
            'date': meeting.date.strftime('%Y-%m-%d'),
            'notes': meeting.notes,
            'attendance': attendance_list
        })
    return jsonify(result)


@app.route('/api/courses/<int:course_id>/meetings', methods=['POST'])
def create_course_meeting(course_id):
    data = request.json
    date = data.get('date')
    notes = data.get('notes', '')
    attendance_ids = data.get('attendance', [])
    # Create meeting
    meeting = CourseMeeting(course_id=course_id, date=datetime.strptime(date, '%Y-%m-%d').date(), notes=notes)
    db.session.add(meeting)
    db.session.commit()
    # Create attendance records for all students enrolled in the course
    enrollments = Enrollment.query.filter_by(course_id=course_id).all()
    for enrollment in enrollments:
        present = enrollment.student_id in attendance_ids
        attendance = Attendance(meeting_id=meeting.id, student_id=enrollment.student_id, present=present)
        db.session.add(attendance)
    db.session.commit()
    return jsonify({'id': meeting.id, 'date': meeting.date.strftime('%Y-%m-%d'), 'notes': meeting.notes}), 201


@app.route('/api/meetings/<int:meeting_id>/attendance', methods=['POST'])
def save_attendance(meeting_id):
    data = request.json  # list of {student_id, present}
    for att in data['attendance']:
        attendance = Attendance.query.filter_by(meeting_id=meeting_id, student_id=att['student_id']).first()
        if attendance:
            attendance.present = att['present']
    db.session.commit()
    return jsonify({'message': 'Attendance saved'}), 200


@app.route('/api/meetings/<int:meeting_id>', methods=['DELETE'])
def delete_meeting(meeting_id):
    try:
        meeting = CourseMeeting.query.get_or_404(meeting_id)
        # remove attendance entries first
        Attendance.query.filter_by(meeting_id=meeting.id).delete()
        db.session.delete(meeting)
        db.session.commit()
        return '', 204
    except Exception as e:
        logger.error(f"Error deleting meeting {meeting_id}: {e}")
        db.session.rollback()
        return jsonify({'error': f'Error deleting meeting: {str(e)}'}), 500


@app.route('/api/calendar/monthly')
def get_monthly_calendar():
    start_date = request.args.get('start_date', datetime.now().strftime('%Y-%m-%d'))
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    # Get the first day of the month
    month_start = start_date.replace(day=1)
    # Get the last day of the month
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1, day=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1, day=1)
    month_end = next_month - timedelta(days=1)

    courses = Course.query.filter(
        Course.start_date <= month_end,
        Course.end_date >= month_start
    ).all()

    calendar_data = []
    for course in courses:
        weekdays = [int(day) for day in course.weekdays.split(',')]
        current_date = max(course.start_date, month_start)
        while current_date <= min(course.end_date, month_end):
            weekday = current_date.weekday()
            sunday_based_weekday = (weekday + 1) % 7
            if sunday_based_weekday in weekdays:
                course_weekdays = [int(day) for day in course.weekdays.split(',')]
                classes_completed = 0
                temp_date = course.start_date
                while temp_date < current_date:
                    temp_weekday = (temp_date.weekday() + 1) % 7
                    if temp_weekday in course_weekdays:
                        classes_completed += 1
                    temp_date += timedelta(days=1)
                classes_remaining = max(0, course.sessions_count - classes_completed)
                enrolled_count = db.session.query(Enrollment).filter_by(course_id=course.id).count()
                calendar_data.append({
                    'id': course.id,
                    'title': course.name,
                    'date': current_date.strftime('%Y-%m-%d'),
                    'time': course.time,
                    'duration': course.duration,
                    'teacher': course.teacher,
                    'color': course.color,
                    'enrolled_count': enrolled_count,
                    'classes_remaining': classes_remaining
                })
            current_date += timedelta(days=1)
    calendar_data.sort(key=lambda x: (x['date'], x['time']))
    return jsonify(calendar_data)


@app.route('/api/students/<int:student_id>', methods=['GET'])
def get_student(student_id):
    student = Student.query.get_or_404(student_id)
    return jsonify(student.to_dict())


@app.route('/student/<int:student_id>')
def student_profile(student_id):
    if not session.get('user'):
        return redirect(url_for('login', next=request.path))
    student = Student.query.get_or_404(student_id)
    # Get all enrollments and courses
    enrollments = Enrollment.query.filter_by(student_id=student_id).all()
    courses = [
        {
            **enrollment.course.to_dict(),
            'enrollment_id': enrollment.id,
            'enrollment_date': enrollment.enrollment_date.strftime('%Y-%m-%d') if enrollment.enrollment_date else None
        }
        for enrollment in enrollments
    ]
    # Get all payments for this student
    payments = Payment.query.filter_by(student_id=student_id).all()
    # Group payments by course and month
    payments_by_course = {}
    for payment in payments:
        course_id = payment.course_id
        if course_id not in payments_by_course:
            payments_by_course[course_id] = []
        payments_by_course[course_id].append({
            'month': payment.month,
            'amount': payment.amount,
            'payment_date': payment.payment_date.strftime('%Y-%m-%d'),
            'course_name': payment.course.name,
            'teacher_name': payment.course.teacher
        })
    return render_template(
        'student.html',
        student=student.to_dict(),
        courses=courses,
        payments_by_course=payments_by_course
    )


@app.route('/api/payments/export', methods=['GET'])
def export_payments_to_excel():
    try:
        period = request.args.get('period', 'all')

        # Build date filter based on period
        if period == 'month':
            start_date = datetime.now().replace(day=1).date()
        elif period == 'quarter':
            current_month = datetime.now().month
            quarter_start_month = ((current_month - 1) // 3) * 3 + 1
            start_date = datetime.now().replace(month=quarter_start_month, day=1).date()
        elif period == 'year':
            start_date = datetime.now().replace(month=1, day=1).date()
        else:  # all
            start_date = None

        # Query payments with optional date filtering
        if start_date:
            payments = Payment.query.filter(Payment.payment_date >= start_date).all()
        else:
            payments = Payment.query.all()

        logger.info(f"Found {len(payments)} payments for export")

        # Check if there are any payments to export
        if not payments:
            logger.info("No payments found for export")
            return jsonify({'error': 'No payments found for the selected period'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"

        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Add export metadata FIRST (rows 1-3)
        ws.cell(row=1, column=1, value="Sports Club Management System - Payment Export").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = Font(
            size=10)
        ws.cell(row=3, column=1,
                value=f"Period: {period.replace('_', ' ').title() if period != 'all' else 'All Time'}").font = Font(
            size=10)

        # Merge metadata cells
        ws.merge_cells('A1:H1')
        ws.merge_cells('A2:H2')
        ws.merge_cells('A3:H3')

        # Style metadata rows
        for row in range(1, 4):
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write header at row 4
        ws.cell(row=4, column=1, value="Invoice Number").font = header_font
        ws.cell(row=4, column=2, value="Payment Date").font = header_font
        ws.cell(row=4, column=3, value="Student Name").font = header_font
        ws.cell(row=4, column=4, value="Course Name").font = header_font
        ws.cell(row=4, column=5, value="Teacher Name").font = header_font
        ws.cell(row=4, column=6, value="Month").font = header_font
        ws.cell(row=4, column=7, value="Amount (₪)").font = header_font
        ws.cell(row=4, column=8, value="Payment Method").font = header_font

        # Apply styles to header
        for row in ws.iter_rows(min_row=4, max_row=4):
            for cell in row:
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

        # Write data
        for i, payment in enumerate(payments, start=5):  # Start from row 5 (after metadata)
            try:
                # Check if relationships exist
                if not payment.student:
                    logger.warning(f"Payment {payment.id} has no student relationship")
                    student_name = "Unknown Student"
                else:
                    student_name = f"{payment.student.first_name} {payment.student.fathers_name}"

                if not payment.course:
                    logger.warning(f"Payment {payment.id} has no course relationship")
                    course_name = "Unknown Course"
                    teacher_name = "Unknown Teacher"
                else:
                    course_name = payment.course.name
                    teacher_name = payment.course.teacher

                ws.cell(row=i, column=1, value=f"INV-{payment.id:06d}")
                try:
                    payment_date = payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else 'Unknown'
                except:
                    payment_date = 'Unknown'
                ws.cell(row=i, column=2, value=payment_date)
                ws.cell(row=i, column=3, value=student_name)
                ws.cell(row=i, column=4, value=course_name)
                ws.cell(row=i, column=5, value=teacher_name)
                ws.cell(row=i, column=6, value=payment.month or 'Unknown')
                ws.cell(row=i, column=7, value=payment.amount or 0)

                # Safely get payment_method - handle case where column might not exist
                try:
                    payment_method = payment.payment_method or 'N/A'
                except AttributeError:
                    # If payment_method attribute doesn't exist, use 'N/A'
                    payment_method = 'N/A'

                ws.cell(row=i, column=8, value=payment_method)

                # Add borders to data cells
                for col in range(1, 9):
                    ws.cell(row=i, column=col).border = border
            except Exception as e:
                logger.error(f"Error processing payment {payment.id}: {str(e)}")
                # Fill with error data
                ws.cell(row=i, column=1, value=f"INV-{payment.id:06d}")
                ws.cell(row=i, column=2, value="Error")
                ws.cell(row=i, column=3, value="Error")
                ws.cell(row=i, column=4, value="Error")
                ws.cell(row=i, column=5, value="Error")
                ws.cell(row=i, column=6, value="Error")
                ws.cell(row=i, column=7, value="Error")
                ws.cell(row=i, column=8, value="Error")

        # Add summary row
        if payments:
            try:
                summary_row = len(payments) + 5  # After data rows (starting from row 5)
                ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
                total_amount = sum(p.amount for p in payments if p.amount is not None)
                ws.cell(row=summary_row, column=7, value=total_amount).font = Font(bold=True)

                # Style the summary row
                for col in range(1, 9):
                    cell = ws.cell(row=summary_row, column=col)
                    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border

                # Add payment method breakdown
                stats_row = summary_row + 2
                ws.cell(row=stats_row, column=1, value="Payment Method Breakdown").font = Font(bold=True, size=12)
                ws.merge_cells(f'A{stats_row}:H{stats_row}')

                # Calculate payment method statistics
                payment_methods = {}
                for payment in payments:
                    try:
                        method = payment.payment_method or 'Unknown'
                    except AttributeError:
                        # If payment_method attribute doesn't exist, use 'Unknown'
                        method = 'Unknown'

                    if method not in payment_methods:
                        payment_methods[method] = {'count': 0, 'amount': 0}
                    payment_methods[method]['count'] += 1
                    payment_methods[method]['amount'] += payment.amount or 0

                # Write payment method statistics
                for i, (method, stats) in enumerate(payment_methods.items()):
                    row = stats_row + 1 + i
                    ws.cell(row=row, column=1, value=method).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=stats['count'])
                    ws.cell(row=row, column=3, value=stats['amount'])
                    if total_amount > 0:
                        percentage = (stats['amount'] / total_amount * 100)
                        ws.cell(row=row, column=4, value=f"{percentage:.1f}%")
                    else:
                        ws.cell(row=row, column=4, value="0.0%")

                    # Style the statistics rows
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                        cell.border = border
            except Exception as e:
                logger.error(f"Error creating summary: {str(e)}")
                # Continue without summary if there's an error

        # Auto-size columns
        try:
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        except Exception as e:
            logger.error(f"Error auto-sizing columns: {str(e)}")
            # Continue with default column widths if auto-sizing fails

        # Create a BytesIO object to hold the Excel file
        try:
            excel_data = BytesIO()
            wb.save(excel_data)
            excel_data.seek(0)
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}")
            return jsonify({'error': f'Error creating Excel file: {str(e)}'}), 500

        # Generate filename with period
        try:
            period_text = period.replace('_', ' ').title() if period != 'all' else 'All Time'
            filename = f'payments_{period_text.replace(" ", "_").lower()}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        except Exception as e:
            logger.error(f"Error generating filename: {str(e)}")
            filename = f'payments_{datetime.now().strftime("%Y%m%d")}.xlsx'

        logger.info(f"Export completed successfully. Filename: {filename}")
        try:
            return send_file(excel_data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             as_attachment=True, download_name=filename)
        except Exception as e:
            logger.error(f"Error sending file: {str(e)}")
            return jsonify({'error': f'Error sending file: {str(e)}'}), 500
    except Exception as e:
        logger.error(f"Error exporting payments to Excel: {str(e)}")
        return jsonify({'error': f'Error exporting payments: {str(e)}'}), 500


if __name__ == '__main__':
    import sys

    if '--reset-db' in sys.argv:
        # Reset database completely
        with app.app_context():
            if reset_database():
                logger.info("Database reset successfully!")
            else:
                logger.error("Failed to reset database!")
        sys.exit(0)
    elif '--populate-test-db' in sys.argv:
        # Use the test database
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///instance/sport_courses_test.db'
        with app.app_context():
            db.drop_all()
            db.create_all()
            # Add Arabic coaches
            coaches = ['أحمد علي', 'سامي يوسف', 'ليلى حسن']
            courses = [
                {'name': 'كرة القدم', 'teacher': coaches[0], 'start_date': '2025-07-01', 'time': '10:00',
                 'sessions_count': 8, 'sessions_per_week': 2, 'weekdays': '0,2', 'color': '#3B82F6'},
                {'name': 'كرة السلة', 'teacher': coaches[1], 'start_date': '2025-07-03', 'time': '12:00',
                 'sessions_count': 10, 'sessions_per_week': 2, 'weekdays': '1,3', 'color': '#F59E42'},
                {'name': 'السباحة', 'teacher': coaches[2], 'start_date': '2025-07-05', 'time': '09:00',
                 'sessions_count': 6, 'sessions_per_week': 1, 'weekdays': '5', 'color': '#10B981'}
            ]
            for c in courses:
                db.session.add(Course(
                    name=c['name'], teacher=c['teacher'],
                    start_date=datetime.strptime(c['start_date'], '%Y-%m-%d').date(),
                    time=c['time'], sessions_count=c['sessions_count'], sessions_per_week=c['sessions_per_week'],
                    weekdays=c['weekdays'], end_date=datetime.strptime(c['start_date'], '%Y-%m-%d').date() + timedelta(
                        weeks=(c['sessions_count'] // c['sessions_per_week'])), color=c['color']
                ))
            db.session.commit()
            # Add Arabic students
            students = [
                {'first_name': 'محمد', 'fathers_name': 'سعيد', 'phone': '0501111111', 'date_of_birth': '2010-01-01',
                 'national_id': '111111111'},
                {'first_name': 'فاطمة', 'fathers_name': 'علي', 'phone': '0502222222', 'date_of_birth': '2011-02-02',
                 'national_id': '222222222'},
                {'first_name': 'يوسف', 'fathers_name': 'حسن', 'phone': '0503333333', 'date_of_birth': '2012-03-03',
                 'national_id': '333333333'},
                {'first_name': 'سارة', 'fathers_name': 'محمود', 'phone': '0504444444', 'date_of_birth': '2013-04-04',
                 'national_id': '444444444'}
            ]
            for s in students:
                db.session.add(Student(
                    first_name=s['first_name'], fathers_name=s['fathers_name'], phone=s['phone'],
                    date_of_birth=datetime.strptime(s['date_of_birth'], '%Y-%m-%d').date(), national_id=s['national_id']
                ))
            db.session.commit()
            # Enroll students in courses
            all_students = Student.query.all()
            all_courses = Course.query.all()
            db.session.add(Enrollment(course_id=all_courses[0].id, student_id=all_students[0].id))
            db.session.add(Enrollment(course_id=all_courses[0].id, student_id=all_students[1].id))
            db.session.add(Enrollment(course_id=all_courses[1].id, student_id=all_students[2].id))
            db.session.add(Enrollment(course_id=all_courses[2].id, student_id=all_students[3].id))
            db.session.commit()
            # Add sample payments
            payments = [
                {'student_id': all_students[0].id, 'course_id': all_courses[0].id, 'month': '2025-07', 'amount': 200,
                 'payment_method': 'cash'},
                {'student_id': all_students[1].id, 'course_id': all_courses[0].id, 'month': '2025-07', 'amount': 200,
                 'payment_method': 'transfer'},
                {'student_id': all_students[2].id, 'course_id': all_courses[1].id, 'month': '2025-07', 'amount': 250,
                 'payment_method': 'check'},
                {'student_id': all_students[3].id, 'course_id': all_courses[2].id, 'month': '2025-07', 'amount': 180,
                 'payment_method': 'cash'},
                {'student_id': all_students[0].id, 'course_id': all_courses[0].id, 'month': '2025-08', 'amount': 200,
                 'payment_method': 'transfer'},
                {'student_id': all_students[1].id, 'course_id': all_courses[0].id, 'month': '2025-08', 'amount': 200,
                 'payment_method': 'cash'},
            ]
            for p in payments:
                db.session.add(Payment(
                    student_id=p['student_id'], course_id=p['course_id'], month=p['month'], amount=p['amount'],
                    payment_method=p['payment_method']
                ))
            db.session.commit()
            logger.info('Test database populated with Arabic data and payments.')
        sys.exit(0)
    
    # Initialize database
    with app.app_context():
        migrate_existing_data()

    # Only run the development server if not using gunicorn
    if not os.getenv('GUNICORN_CMD_ARGS'):
        # Get port from environment variable or use default
        port = int(os.environ.get('PORT', 5000))
        debug = os.environ.get('FLASK_ENV') == 'development'
        app.run(debug=debug, host='0.0.0.0', port=port)

# ------------------- TESTS -------------------
import unittest


class APITestCase(unittest.TestCase):
    def setUp(self):
        app.config['TESTING'] = True
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
        self.app = app.test_client()
        with app.app_context():
            db.create_all()

    def tearDown(self):
        with app.app_context():
            db.drop_all()

    def test_create_student_valid(self):
        resp = self.app.post('/api/students', json={
            'first_name': 'Ali',
            'fathers_name': 'Hassan',
            'phone': '0501234567',
            'date_of_birth': '2010-01-01',
            'national_id': '123456789'
        })
        self.assertEqual(resp.status_code, 201)
        data = resp.get_json()
        self.assertEqual(data['first_name'], 'Ali')

    def test_create_student_duplicate_phone(self):
        self.test_create_student_valid()
        resp = self.app.post('/api/students', json={
            'first_name': 'Sara',
            'fathers_name': 'Omar',
            'phone': '0501234567',  # duplicate
            'date_of_birth': '2011-01-01',
            'national_id': '987654321'
        })
        self.assertEqual(resp.status_code, 409)

    def test_create_course_valid(self):
        resp = self.app.post('/api/courses', json={
            'name': 'Football',
            'teacher': 'Coach Sam',
            'start_date': '2025-01-01',
            'time': '10:00',
            'sessions_count': 8,
            'sessions_per_week': 2,
            'weekdays': '0,2',
            'color': '#3B82F6'
        })
        self.assertEqual(resp.status_code, 201)
        data = resp.get_json()
        self.assertEqual(data['name'], 'Football')

    def test_create_course_duplicate(self):
        self.test_create_course_valid()
        resp = self.app.post('/api/courses', json={
            'name': 'Football',
            'teacher': 'Coach Sam',
            'start_date': '2025-01-01',
            'time': '10:00',
            'sessions_count': 8,
            'sessions_per_week': 2,
            'weekdays': '0,2',
            'color': '#3B82F6'
        })
        self.assertEqual(resp.status_code, 409)

    def test_enroll_student(self):
        self.test_create_student_valid()
        self.test_create_course_valid()
        # Get IDs
        students = self.app.get('/api/students').get_json()
        courses = self.app.get('/api/courses').get_json()
        student_id = students[0]['id']
        course_id = courses[0]['id']
        resp = self.app.post('/api/enrollments', json={
            'student_id': student_id,
            'course_id': course_id
        })
        self.assertEqual(resp.status_code, 201)

    def test_payment_and_invoice(self):
        self.test_enroll_student()
        students = self.app.get('/api/students').get_json()
        courses = self.app.get('/api/courses').get_json()
        student_id = students[0]['id']
        course_id = courses[0]['id']
        # Add payment
        resp = self.app.post('/api/payments', json={
            'student_id': student_id,
            'course_id': course_id,
            'month': '2025-01',
            'amount': 200
        })
        self.assertEqual(resp.status_code, 201)
        payment_id = resp.get_json()['id']
        # Get invoice
        invoice = self.app.get(f'/api/invoice/{payment_id}')
        self.assertEqual(invoice.status_code, 200)
        self.assertIn('invoice_number', invoice.get_json())

    def test_meeting_and_attendance(self):
        self.test_enroll_student()
        students = self.app.get('/api/students').get_json()
        courses = self.app.get('/api/courses').get_json()
        student_id = students[0]['id']
        course_id = courses[0]['id']
        # Add meeting
        resp = self.app.post(f'/api/courses/{course_id}/meetings', json={
            'date': '2025-01-10',
            'notes': 'First session'
        })
        self.assertEqual(resp.status_code, 201)
        meeting_id = resp.get_json()['id']
        # Save attendance
        resp2 = self.app.post(f'/api/meetings/{meeting_id}/attendance', json={
            'attendance': [
                {'student_id': student_id, 'present': True}
            ]
        })
        self.assertEqual(resp2.status_code, 200)
        self.assertIn('Attendance saved', resp2.get_json()['message'])


if __name__ == '__main__':
    import sys

    if 'test' in sys.argv:
        unittest.main(argv=['first-arg-is-ignored'], exit=False)
_schema_checked = False

@app.before_request
def ensure_schema_on_first_request():
    """Ensure required tables exist; runs once per process.
    Compatible with Flask versions that don't expose before_first_request.
    """
    global _schema_checked
    if _schema_checked:
        return
    try:
        inspector = db.inspect(db.engine)
        if 'coaches' not in inspector.get_table_names():
            # Only create the coaches table if it's missing; avoid heavy migrations here
            Coach.__table__.create(db.engine)
            logger.info("Created 'coaches' table")
    except Exception as e:
        logger.error(f"Failed to ensure coaches table exists: {e}")
    finally:
        _schema_checked = True
